#
# Copyright 2013 Tail-f Systems
#
#(C) 2017 Tail-f Systems

import json
from nso_api import NsoSession as nso
import time
from datetime import datetime, time
from openpyxl import load_workbook

columns = {'vrf' : 0,
           'description' : 1,
           'device' : 2,
           'route-distinguisher': 3,
           'import-route-policy' : 4,
           'export-route-policy' : 5,
           'import-route-target' : 6,
           'export-route-target' : 7,
           'max-prefix-limit'    : 8,
           'max-threshold'       : 9}

##
# Path for vrf spreadsheet
##
filepath = '/Users/dan/Desktop/vrf-onboarding/'

class vrf:

    def __init__(self, session):
      self.nso = session
      return

    def onboardSyncFrom(self):
       """Sync from Device node(s) and Service node(s)"""
       dNodes = [{'node' :'nso-1', 'address' : '127.0.0.1', 'port' :8081},
              {'node':'nso-2', 'address' : '127.0.0.1', 'port':8082}]
       ###
       # Device nodes have to be sync'd first
       ###
       for nd in dNodes:
          # open a connection to the device node
          print "Syncing node %s" % nd['node']
          nsoc = nso(nd['address'], None, self.nso.user, self.nso.passwd, port=nd['port'], debug=False)
          status = nsoc.login(self.nso.user, self.nso.passwd)
          if not status:
             print "\t\tFailed login %s" % nd['node']
             return False
          nsoc.getNewTrans('read')
          nsoc.action('/ncs:devices/sync-from')
          nsoc.logout()
       ##
       # Just sync the current node
       ## 
       self.nso.action('/ncs:devices/sync-from')

       return True

    def onboardCreate(self, vrfd):
        """First step is to add the VRF instances to NSO"""
        results = {}
        start = datetime.now()

        for vrfName, vrf in vrfd.items():
            
            # Create the path
            path = '/ncs:services/svrf:vrf{' + vrfName + '}'  
      
            if self.nso.exists(path) == False :
              result = self.nso.create(path)
            else:
              vrf['state'] = 'Created'
              print "\tVRF service instance [%s] already exists" % vrfName 
            # Set top level serivce parameters 
            self.nso.set(path + '/description', vrf['description'])

            ##
            # Configure each device
            ##
            for dev in vrf['devices']:
              epath = path + '/devices{' + dev['name'] + '}'
              if self.nso.exists(epath) == False:
                self.nso.create(epath)
              
              if 'route-distinguisher' in dev.keys():
                self.nso.set(epath + '/route-distinguisher', dev['route-distinguisher'])
              if 'import-route-policy' in dev.keys():
                self.nso.set(epath + '/import-route-policy', dev['import-route-policy'])
              if 'max-prefix-limit' in dev.keys():
                self.nso.set(epath + '/max-prefix-limit', str(dev['max-prefix-limit']))
              if 'max-threshold' in dev.keys():
                self.nso.set(epath + '/max-prefix-threshold', str(dev['max-threshold']))
              if 'export-route-policy' in dev.keys():
                self.nso.set(epath + '/export-route-policy', dev['export-route-policy'])
              if 'import-route-target' in dev.keys():
                for target in dev['import-route-target'].split(","):
                  targ = epath + '/import-route-target{' + target + '}'
                  if self.nso.exists(targ) == False:
                    self.nso.create(targ)
              if 'export-route-target' in dev.keys():
                 for target in dev['export-route-target'].split(","):
                    targ = epath + '/export-route-target{' + target + '}'
                    if self.nso.exists(targ)  == False:
                      self.nso.create(targ)
              
            ## 
            # Commit the changes use no-deploy
            ##
            self.nso.validateCommit()  
            result = self.nso.commit(flags=['no-networking'])
            if result == True:
              vrf['state'] = 'Created'
        
        results['status'] = 'Create operation(s) completed'
        finish = datetime.now()
        elapsed = finish - start
        results['message'] = "\nStart Time %s\nFinish Time %s\nElapsed Time : %s" % (str(start), str(finish),str(elapsed))
        return results

    def getConfiguration(self, vrfName):
        """"Execute a get-modifications with dry run """
        
        path = '/ncs:services/svrf:vrf{' + vrfName + '}'
        res = self.nso.get(path)

        if 'result' in res.keys() and  'config' in res['result'].keys():
           if 'config' in res['result'].keys():
             return res['result']['config']
        return "No NSO configuration available"

    def getModifications(self, vrfName):
        """"Execute a get-modifications"""
        print "Execute get-modifications"
        path = '/ncs:services/svrf:vrf{' + vrfName + '}/get-modifications'
        res = self.nso.solo_action(path)
 
        if 'result' not in res:
           return "\ncli  {\n}"

        result = "\ncli  {\n"
        for entry in res['result']:
  
          if entry['name'] == 'cli/local-node/data':
             result += "     local-node {\n         data "
             result += entry['value'].replace("\n", "\n         ")
             result += "\n    }"
          if entry['name'] == 'cli/lsa-service/service-id':
             result += "\n    lsa-service {\n        service-id  %s\n        data  " % entry['value']
  
          if entry['name'] == 'cli/lsa-service/data':
             result += entry['value'].replace("\n", "\n         ")
             result += "\n     }"
        result += "\n}"

        return result

    def dryReconcile(self, vrfName):
        """"Execute a re-deploy with dry run """

        path = '/ncs:services/svrf:vrf{' + vrfName + '}/re-deploy'
        result = self.nso.redeployReconcile(path, dryrun=True)

        if not result:
             return "Service is currently reconciled"
        if 'result' in result:
           ##
           # Format the output
           ##
           res = self.nso.formatDryRun(result, 'cli')
           return res
      
        return "Reconcile operation failed"

    def forceReconcile(self, vrfd, vrfName):
        """"Execute a re-deploy without dry run """

        path = '/ncs:services/svrf:vrf{' + vrfName + '}/re-deploy'
        result = self.nso.redeployReconcile(path, dryrun=False)
        if 'result' in result and result['result'] == True:
             print "Update vrf state"
             ##
             # Mark the VRF serfice as reconciled
             ##
             print vrfd
             vrfd[vrfName]['state'] = 'Reconciled'
             vrfd[vrfName]['modfied'] = True
             return "Successfully reconciled %s" % vrfName

        return "Force re-deploy reconcile failed"

    def onboardReconcile(self, vrfd):
        """2nd step is to attempt the VRF service reconcile"""
        
        results = {}
        start = datetime.now()
        for vrfName, vrf in vrfd.items():

            # Create the path
            path = '/ncs:services/svrf:vrf{' + vrfName + '}/re-deploy'
            print "Reconcile: %s" % path
            ###
            # First dry-run the reconcile to see if the reconcile will
            # require network changes
            ###
            result = self.nso.redeployReconcile(path, dryrun=True)
            
            if not result:
               ##
               # If the reconcile doesn't require a network change
               # perform the reconcile
               ##
               result = self.nso.redeployReconcile(path, dryrun=False)
               vrf['state']= 'Reconciled'

            else:
               vrf['state'] = 'Failed'

        results['status'] = 'Reconcile operation(s) completed'
        finish = datetime.now()
        elapsed = finish - start
        results['message'] = "\nStart Time %s\nFinish Time %s\nElapsed Time : %s" % (str(start), str(finish),str(elapsed))
        return results

    def onboardNetReconcile(self, vrfd):
          """2nd step is to attempt the VRF service reconcile"""
        
          results = {}
          start = datetime.now()
          for vrfName, vrf in vrfd.items():

            if vrf['state'] != 'Failed':
               continue

            ##
            # Issue the network reconcile for each service which has not been
            # successfully reconciled
            ##
            path = '/ncs:services/svrf:vrf{' + vrfName + '}/oob-reconcile'
            print '\t\tReconcile [%s] network deployment w/service conifguration' % vrfName
            data = self.nso.solo_action(path)

          ###
          # Now once again try to reconcile the service
          ###
          for vrfName, vrf in vrfd.items():

             if vrf['state'] != 'Failed':
               continue

             # Create the path
             path = '/ncs:services/svrf:vrf{' + vrfName + '}/re-deploy'
             print "Reconcile: %s" % path

             result = self.nso.redeployReconcile(path, dryrun=True)
            
             if not result:
               ##
               # If the reconcile doesn't require a network change
               # perform the reconcile
               ##
               result = self.nso.redeployReconcile(path, dryrun=False)
               vrf['state']= 'Reconciled'
               vrf['modified'] = True
             else:
               vrf['state'] = 'Failed'

          results['status'] = 'Network Reconcile operation(s) completed'
          finish = datetime.now()
          elapsed = finish - start
          results['message'] = "\nStart Time %s\nFinish Time %s\nElapsed Time : %s" % (str(start), str(finish),str(elapsed))
          return results

    def onboardNetDiscover(self, vrfd):
        """Discover the VRF service(s)"""

        results = {}
        start = datetime.now()

        ###
        # Issue the network discovery command
        ###
        data = self.nso.solo_action('/svrf:cmd/vrf-discover')
        try :
          result = data['result'] 
          vrfstr = result[1]['value']
          vrfs = vrfstr[1:-1].split()
        except:
          vrfstr = ''
          vrfs = []
        
        ###
        # For each VRF returned issue the re-deploy reconcile
        ###
        for vrf in vrfs:
              path = '/ncs:services/svrf:vrf{' + vrf + '}/re-deploy'
              result = self.nso.redeployReconcile(path, dryrun=True)
             
              if not result:
                ##
                # If the reconcile doesn't require a network change
                # perform the reconcile
                ##
                result = self.nso.redeployReconcile(path, dryrun=False)
                print "\t\tSuccessfully Reconciled service %s" % vrf
         
                query_info = ['name', 'description']
                sxpath = "/ncs:services/svrf:vrf[name=\'%s']" % vrf
                data = self.nso.simple_query('', sxpath, query_info)
                vrfa = [ dict(zip(query_info, values)) for values in data['result']['results']]

                for v in vrfa:
   
                  query_info = ['name', 'route-distinguisher', 'import-route-policy', 'export-route-policy','import-route-target', 'export-route-target', 'max-prefix-limit', 'max-prefix-threshold']
                  data = self.nso.simple_query('',sxpath + '/devices', query_info)
                  devs = [ dict(zip(query_info, values)) for values in data['result']['results']]
                  v['devices'] = devs
                  for dev in devs:
                    xpath = sxpath + "/devices[\'%s\']/import-route-target" % dev['name']      
                    query_info = ['name']
                    data = self.nso.simple_query('', xpath, query_info)
                    imps = [ dict(zip(query_info, values)) for values in data['result']['results']]

                    dev['import-route-target'] = imps
                    xpath = sxpath + "/devices[\'%s\']/export-route-target" % dev['name']      
                    query_info = ['name']
                    data = self.nso.simple_query('', xpath, query_info)
                    exps = [ dict(zip(query_info, values)) for values in data['result']['results']]      
                    dev['export-route-target'] = exps
            
                  if vrfd.has_key(vrf):
                    vrfd[vrf]['modified'] = True
                  else:
                    ###
                    # Create an inventory entry if one doesn't exist
                    ###
                    vrfd[vrf] = {}
                    vrfd[vrf]['name'] = vrf
                    vrfd[vrf]['description'] = v['description']
                    vrfd[vrf]['state'] = 'Discovered'
                    vrfd[vrf]['devices'] = []
                    ds = vrfd[vrf]['devices']
                    devlist = []
                    for dev in v['devices']:
                       if dev['name'] not in devlist:
                         devlist.append(dev['name'])
                       entry = {}
                       entry['name'] = dev['name']
                       entry['route-distinguisher'] = dev['route-distinguisher']
                       entry['import-route-policy'] = dev['import-route-policy']
                       entry['export-route-policy'] = dev['export-route-policy']
                       entry['max-prefix-limit'] = dev['max-prefix-limit']
                       entry['max-threshold'] = dev['max-prefix-threshold']
                       exportRT = ''
                       for expt in dev['export-route-target']:
                          exportRT = exportRT + ',' + expt['name']
                       entry['export-route-target'] = exportRT[1:]
                       importRT = ''
                       for expt in dev['import-route-target']:
                          importRT = importRT + ',' + expt['name']
                       entry['import-route-target'] = importRT[1:]
                       vrfd[vrf]['devices'].append(entry)
                       vrfd[vrf]['route-distinguisher'] = dev['route-distinguisher']
                    print devlist
                    vrfd[vrf]['device-list'] = ",".join(devlist )
      
        results['status'] = 'Network Discover operation(s) completed\n Discovered VRFS %s' % vrfstr
        finish = datetime.now()
        elapsed = finish - start
        results['message'] = "\nStart Time %s\nFinish Time %s\nElapsed Time : %s" % (str(start), str(finish),str(elapsed))
        return results


    def getVrfOnboardStats(self, vrfd):
        stats = {}
        stats['Discovered'] = 0
        stats['Failed'] = 0
        stats['Reconciled'] = 0
        stats['Unknown'] = 0
        stats['Created'] = 0
        stats['Total'] = 0

        for vrf,v in vrfd.items():
          stats[v['state']] += 1

        stats['Total'] = stats['Discovered'] + stats['Failed']
        stats['Total'] += stats['Reconciled'] + stats['Unknown'] 
        stats['Total'] += stats['Created']
        return stats
     
    def writeVrfData(self, vrfd, sheet, file):
        """ Write VRF Reconcile information into EXCEL Workbook """

        print "Write results[%s] sheet[%s]" % (file, sheet)
        wb = load_workbook(filepath + file)
        services = vrfd['services']
        if sheet in wb.sheetnames:
            result = wb[sheet]
        else:
            result = wb.create_sheet(sheet)
            result.title = sheet
        
        result.cell(row=1, column=1).value = 'vrf'
        result.cell(row=1, column=2).value = 'description'
        result.cell(row=1, column=3).value = 'device'
        result.cell(row=1, column=4).value = 'route-distinguisher'
        result.cell(row=1, column=5).value = 'import-route-policy'
        result.cell(row=1, column=6).value = 'export-route-policy'
        result.cell(row=1, column=7).value = 'import-route-target'
        result.cell(row=1, column=8).value = 'export-route-target'
        result.cell(row=1, column=9).value = 'max-prefix-limit'
        result.cell(row=1, column=10).value = 'max-threshold'
        result.cell(row=1, column=11).value = 'state'
        result.cell(row=1, column=12).value = 'modified'
   
        row = 2
        for srvName, srv in services.items():
           print '\tService [%s] state [%s]' % (srvName, srv['state'])    
           for dev in srv['devices']:  
              result.cell(row=row, column=1).value = srvName
              result.cell(row=row, column=2).value = srv['description']
              result.cell(row=row, column=3).value = dev['name']
              result.cell(row=row, column=4).value = dev['route-distinguisher']
              result.cell(row=row, column=5).value = dev['import-route-policy']
              result.cell(row=row, column=6).value = dev['export-route-policy']
              result.cell(row=row, column=7).value = dev['import-route-target']
              result.cell(row=row, column=8).value = dev['export-route-target']
              if 'max-prefix-limit' in dev.keys():
                result.cell(row=row, column=9).value = dev['max-prefix-limit']
              if 'max-threshold' in dev.keys():
                result.cell(row=row, column=10).value = dev['max-threshold']
              result.cell(row=row, column=11).value = srv['state']
              if 'modified' in srv.keys():
                result.cell(row=row, column=12).value = srv['modified']
              row += 1

        print "\t\tSaving file [%s]" % file
        wb.save(filename = filepath + file)   
           
        return True

    def loadVrfData(self, vrfd):

        ##
        # Don't try to read file data if the filename
        # hasn't been specified
        ##
        if vrfd['file'] == 'None':
          vrfd['services'] = {}
          vrfd['statistcis'] = {}
          return {}
        
        if vrfd['services']:
           return vrfd['services']

        services = vrfd['services'] = {}
        wb = load_workbook(filepath + vrfd['file'])
        ws = wb[vrfd['sheet']]
        serviceData = False
        startCol = 0
        endCol = 0
        startRow = 0
        serviceNames = {}

        for row in range(1, 10):
          for col in range  (1, 20):
            rcell = ws.cell(row=row, column=col)
            if rcell.value == 'vrf':
               startCol = col
               startRow = row + 1
               serviceData = True
            if serviceData == True:
               if rcell.value:
                 serviceNames[str(col)] = rcell.value
               else:
                  endCol = col
                  break
          if serviceData == True:
            break;      

        for row in range(startRow, 10000):
            rcell = ws.cell(row=row, column=startCol)
            ##
            # Grab the service instance entry or create one
            ##
            if not rcell.value:
               break;
            vrf = str(rcell.value)
            if vrf in services:
               srv = services[vrf]
               
            else:
               srv = services[vrf] = {}
               srv['devices'] = []
               srv['state'] = 'Unknown'
               srv['device-list'] = None

            devices = srv['devices']
           
            srv['name'] = str(ws.cell(row=row, 
                                      column=startCol+columns['vrf']).value)
            srv['description'] = str(ws.cell(row=row, 
                                    column=startCol+columns['description']).value)
            srv['route-distinguisher'] = str(ws.cell(row=row, 
                                    column=startCol+columns['route-distinguisher']).value)
            ##
            # Fill in each device for the VRF
            ##            
            if not ws.cell(row=row, column=startCol+columns['device']).value:
               continue
            
            devName = ws.cell(row=row, column=startCol+columns['device']).value

            if srv['device-list']:
               srv['device-list'] = srv['device-list'] + "," + devName
            else:
               srv['device-list'] = devName

            dev = {}
            dev['name'] = devName
            if ws.cell(row=row, column=startCol+columns['route-distinguisher']).value:
                dev['route-distinguisher'] = ws.cell(row=row, 
                   column=startCol+columns['route-distinguisher']).value 
            if ws.cell(row=row, column=startCol+columns['import-route-policy']).value:
                dev['import-route-policy'] = ws.cell(row=row, 
                   column=startCol+columns['import-route-policy']).value            
            if ws.cell(row=row, column=startCol+columns['export-route-policy']).value:
                dev['export-route-policy'] = ws.cell(row=row, 
                  column=startCol+columns['export-route-policy']).value
            if ws.cell(row=row, column=startCol+columns['import-route-target']).value:
                dev['import-route-target'] = ws.cell(row=row, 
                  column=startCol+columns['import-route-target']).value 
            if ws.cell(row=row, column=startCol+columns['export-route-target']).value:
                dev['export-route-target'] = ws.cell(row=row, 
                  column=startCol+columns['export-route-target']).value
            if ws.cell(row=row, column=startCol+columns['max-prefix-limit']).value:
                dev['max-prefix-limit'] = ws.cell(row=row, 
                  column=startCol+columns['max-prefix-limit']).value
            if ws.cell(row=row, column=startCol+columns['max-threshold']).value:
                dev['max-threshold'] = ws.cell(row=row, 
                  column=startCol+columns['max-threshold']).value
            devices.append(dev)

        return services


    def getVrfs(self):
        query_info = ['name', 'description']
        
        data = self.nso.simple_query('/ncs:services', 'vrf', query_info)
        vrfsd = data['result']['results']
        vrfs = [ dict(zip(query_info, values)) for values in vrfsd]
        
        query_info = ['name', 'route-distinguisher', 'import-route-target', 'export-route-target',
                      'import-route-policy', 'export-route-policy', 
                      'max-prefix-limit', 'max-prefix-threshold']
        query_info.append('device')
          
        return vrfs

    def addVrf(self, vrfParams):
        
       addType = vrfParams['type']
       ###
       # Add a VRF or a VRF device
       ###
       if addType == 'vrf':

           vrf = vrfParams['name']
           path = '/ncs:services/vrf:vrf{' + vrf + '}'
       
           if self.nso.exists(path) == False :
             result = self.nso.create(path)

           self.nso.set(path + '/route-distinguisher', vpnParams['rd'])
                            
       return True

    def vrfDel(self, vrfParams):

       path = '/ncs:services/vrf:vrf{' + vrfParams['vrf'] + '}'
      
       if vpnParams['type'] == 'device':
          path += '/device{' + vrfParams['device'] + '}'

       if self.nso.exists(path) == False :
          return('Failed: device does not exists')       

       return (self.nso.delete(path))
      

